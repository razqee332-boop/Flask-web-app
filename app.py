from flask import Flask, render_template, request, send_file
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import requests
from openpyxl.drawing.image import Image as ExcelImage

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
IMAGE_FOLDER = "images"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(IMAGE_FOLDER, exist_ok=True)


def login_to_website(username, password):
    """Login using Selenium and return driver object."""
    driver = webdriver.Chrome()

    driver.get("https://yourcompanywebsite.com/login")  # <-- EDIT THIS

    # ENTER USERNAME + PASSWORD (EDIT SELECTORS)
    driver.find_element(By.ID, "username").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "loginBtn").click()

    time.sleep(3)
    return driver


def fetch_first_image(driver, barcode):
    """Search barcode and return the first image URL."""
    try:
        # FIND SEARCH BOX (EDIT THIS SELECTOR)
        search_box = driver.find_element(By.ID, "searchInput")
        search_box.clear()
        search_box.send_keys(barcode)
        search_box.send_keys(Keys.ENTER)

        time.sleep(2)

        # FIND FIRST PRODUCT IMAGE (EDIT SELECTOR TO MATCH YOUR SITE)
        img_el = driver.find_element(By.CSS_SELECTOR, ".product-image")
        img_url = img_el.get_attribute("src")

        return img_url

    except:
        return None


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        excel_file = request.files["file"]
        username = request.form["username"]
        password = request.form["password"]

        filepath = os.path.join(UPLOAD_FOLDER, excel_file.filename)
        excel_file.save(filepath)

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Login to website
        driver = login_to_website(username, password)

        # Process barcodes
        for row in range(2, ws.max_row + 1):
            barcode = ws.cell(row=row, column=1).value

            img_url = fetch_first_image(driver, barcode)

            if img_url:
                img_data = requests.get(img_url).content
                img_path = os.path.join(IMAGE_FOLDER, f"{barcode}.jpg")
                with open(img_path, "wb") as f:
                    f.write(img_data)

                # Insert image into Excel
                excel_img = ExcelImage(img_path)
                excel_img.width = 120
                excel_img.height = 120

                ws.add_image(excel_img, f"B{row}")

            ws.cell(row=row, column=3).value = img_url or "Not found"

        driver.quit()

        # Save output Excel
        output_file = os.path.join(UPLOAD_FOLDER, "output_with_images.xlsx")
        wb.save(output_file)

        return send_file(output_file, as_attachment=True)

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
